from os import listdir, path, mkdir
from os.path import basename, dirname, join, exists
import xlrd
import openpyxl

def get_list_of_input_files(dir_with_invoices):
    '''Function for getting a list of files with archives'''

    input_files = listdir(dir_with_invoices)
    invoice_name = 'Архив счетов-фактур'
    input_files_temp = []
    for file in input_files:
        if file.startswith(invoice_name):
            input_files_temp.append(file)
    input_files = input_files_temp
    return input_files

def get_float_from_cell(cell):
    '''function to extract floating point value from cell'''

    cell_value = str(cell.value)
    if ',' in cell_value:
        result = ''
        for char in cell_value:
            if char.isdigit():
                result += char
            elif char == ',':
                result += '.'
        if not result.endswith('.'):
            return float(result)
        else:
            return None
    else:
        try:
            return float(cell_value)
        except:
            return None

def calculate_profit(input_file, include_groups = False):
    '''The function receives the path to the invoice archive file,
    calculates the indicators and returns a dictionary with values'''

    result = {}
    # Extracting date from file name
    can_get_date = (
        input_file.find(' с ') != -1,
        input_file.find(' по ') != -1,
        )
    result['Invoice_date'] = None
    if all(can_get_date):
        result['Invoice_date'] = input_file[input_file.find(' с ')+1:input_file.find(' по ')+14]
    
    # Opening the file and getting the table from the first sheet
    table = xlrd.open_workbook(input_file).sheet_by_index(0)
    
    # Getting columns with name, sum, sales volumes and optionally groups
    products = table.col(0)[3:-3]
    products_sum = table.col(3)[3:-3]
    product_sales_volumes = table.col(2)[3:-3]
    product_groups = None
    if include_groups:
        product_groups = table.col(1)[3:-3]
    
    result_products = {}
    total_revenue = 0

    for index in range(len(products)):
        product = products[index].value
        product_sum = get_float_from_cell(products_sum[index])
        product_sales_volume = get_float_from_cell(product_sales_volumes[index])
        
        not_valid_data = (
            product.startswith('Наименование'),
            product.startswith('№'),
            product.startswith('Итого:'),
            product.startswith('В том числе НДС'),
            (product) == '',
            not product_sum,
            not product_sales_volume
            )
        if any(not_valid_data):
            continue
        
        total_revenue += product_sum

        if result_products.get(product):
            result_products[product]['product_sum'] += product_sum
            result_products[product]['product_sales_volume'] += product_sales_volume
        else:
            result_products[product] = {
            'product_sum' : product_sum,
            'product_sales_volume' : product_sales_volume}
        
        if include_groups:
            result_products[product]['group'] = product_groups[index].value
        else:
            result_products[product]['group'] = None

    result['result_products'] = result_products
    result['total_revenue'] = total_revenue
    result['include_groups'] = include_groups
    return result

def write_results(results_list):
    '''The function takes a list of results and writes them to the appropriate files'''

    bd = openpyxl.styles.Side(style='thin', color="000000")
    all_borders = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
    allignment_center = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')
    bold_font = openpyxl.styles.Font(bold=True)

    counter = 1
    for result in results_list:
        result_products = result['result_products']

        wb = openpyxl.Workbook()
        ws = wb['Sheet']

        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
        
        identificator = counter
        if result.get('Invoice_date'):
            identificator = result['Invoice_date']

        ws['A1'] = 'Общая выручка: '
        ws['A1'].font = bold_font
        ws['A1'].border = all_borders

        ws['A2'] = 'Приблизительная сумма прибыли при наценке 5%: '
        ws['A2'].font = bold_font
        ws['A2'].border = all_borders

        total_revenue = result['total_revenue']
        ws['C1'] = total_revenue
        ws['C1'].alignment = allignment_center
        ws['C1'].border = all_borders
        ws['C1'].number_format = '0.00'

        ws['C2'] = (total_revenue / 1.05 / 1.2) * 0.05
        ws['C2'].number_format = '0.00'
        ws['C2'].alignment = allignment_center
        ws['C2'].border = all_borders

        ws['A3'] = 'Приблизительная сумма прибыли при наценке (укажите процент): '
        ws['A3'].font = bold_font
        ws['A3'].border = all_borders

        ws['B3'] = 0
        ws['B3'].number_format = '0.00%'
        ws['B3'].alignment = allignment_center
        ws['B1'].border = all_borders
        ws['B2'].border = all_borders
        ws['B3'].border = all_borders

        ws['C3'] = '=C1/1.2/(1+B3)*B3'
        ws['C3'].number_format = '0.00'
        ws['C3'].alignment = allignment_center
        ws['C3'].border = all_borders

        ws['A5'] = 'Товар'
        ws['A5'].font = bold_font
        ws['A5'].alignment = allignment_center
        ws['A5'].border = all_borders

        ws['B5'] = 'Продано единиц'
        ws['B5'].font = bold_font
        ws['B5'].alignment = allignment_center
        ws['B5'].border = all_borders

        ws['C5'] = f'Сумма продаж за период {identificator} по товару'
        ws['C5'].font = bold_font
        ws['C5'].alignment = allignment_center
        ws['C5'].border = all_borders

        if result['include_groups']:
            ws['D5'] = "Группа товара"
            ws['D5'].font = bold_font
            ws['D5'].alignment = allignment_center
            ws['D5'].border = all_borders
        
        result_products_sorted = list(result_products.keys())
        result_products_sorted.sort(key=lambda product: result_products[product]['product_sum'], reverse=True)

        for index, product in enumerate(result_products_sorted):
            ws['A'+str(index+6)] = product
            ws['A'+str(index+6)].border = all_borders

            ws['B'+str(index+6)] = result_products[product]['product_sales_volume']
            ws['B'+str(index+6)].alignment = allignment_center
            ws['B'+str(index+6)].border = all_borders

            ws['C'+str(index+6)] = result_products[product]['product_sum']
            ws['C'+str(index+6)].number_format = '0.00'
            ws['C'+str(index+6)].alignment = allignment_center
            ws['C'+str(index+6)].border = all_borders

            if result['include_groups']:
                ws['D'+str(index+6)] = result_products[product]['group']
                ws['D'+str(index+6)].alignment = allignment_center
                ws['D'+str(index+6)].border = all_borders
        
        output_file = f'Отчёт по продажам {identificator}.xlsx'
        output_file = join('Отчёты', output_file)
        wb.save(output_file)
        counter += 1


def main():
    dir_with_invoices = join(dirname(__file__), 'Архивы счетов-фактур')
    if not exists(dir_with_invoices):
        mkdir(dir_with_invoices)
        print('Создана папка для архивов счетов-фактур.\n'+
            'Перенесите архивы в папку "Архивы счетов-фактур",\n' +
            'чтобы сформировать отчёты о продажах.')
        input('Для продолжения нажмите Enter')
    
    empty_dir = True
    while empty_dir:
        empty_dir = not bool(listdir(dir_with_invoices))
        if empty_dir:
            input('В папке "Архивы счетов-фактур" не найдено ни одного архива.\n' + 
                'Перенесите архивы в папку "Архивы счетов-фактур", чтобы сформировать отчёты о продажах,\n' +
                'и нажмите Enter')

    print('Получение списка архивов счетов-фактур...')
    input_files = get_list_of_input_files(dir_with_invoices)

    results = []
    counter = 1
    number_of_files = len(input_files)

    for file in input_files:
        print(f'Рассчитываем {counter} отчёт из {number_of_files}...')
        input_file_path = join(dir_with_invoices, file)
        results.append(calculate_profit(input_file_path, include_groups=True))
        counter += 1

    print('Сохраняем отчёты в Excel...')
    dir_to_save_results = join(dirname(__file__), 'Отчёты')
    if not exists(dir_to_save_results):
        mkdir(dir_to_save_results)

    write_results(results)
    print('Отчёты о продажах сохранены в папку "Отчёты".')
    input('Для завершения нажмите Enter')


if __name__=='__main__':
    main()
